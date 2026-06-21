#!/usr/bin/env python3
"""Preprocess the source workbook into frontend-friendly JSON files."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "roman-wb-valentin" / "2026-06-17_roman-wb.xlsx"
DEFAULT_OUTPUT_DIR = ROOT / "data" / "processed"

DASH_VALUES = {"", "-", "–", "—", "―"}

GLOSSARY_SHEET = "GLOSSARY"
ABBREVIATION_SHEETS = ("abbrs-gram", "abbrs-lang", "abbrs-lex")
PARADIGM_SHEETS = (
    "ADJ-DECL",
    "F-DECL",
    "M-DECL",
    "MF-DECL",
    "V-CONJG",
    "V-EXIST",
)

EXPECTED_GLOSSARY_COLUMNS = (
    "ROMAN INT",
    "ROMAN DEU",
    "Composition INT",
    "Composition DEU",
    "Variation INT",
    "Variation DEU",
    "Reconstruction INT",
    "Reconstruction DEU",
    "Source-1",
    "Source-2 INT",
    "Source-2 DEU",
    "Base INT",
    "Base DEU",
    "Word class 1",
    "Word class 2",
    "Flexion 1",
    "Flexion 2 INT",
    "Flexion 2 DEU",
    "Flexion 3 INT",
    "Flexion 3 DEU",
    "Paradigm",
    "Domain",
    "DEUTSCH 01",
    "DEUTSCH 02",
    "DEUTSCH 03",
    "DEUTSCH 04",
    "DEUTSCH 05",
    "DEUTSCH 06",
    "DEUTSCH 07",
    "DEUTSCH 08",
    "DEUTSCH 09",
    "DEUTSCH 10",
    "ENGLISH 01",
    "ENGLISH 02",
    "ENGLISH 03",
    "ENGLISH 04",
    "ENGLISH 05",
    "ENGLISH 06",
    "ENGLISH 07",
    "ENGLISH 08",
    "ENGLISH 09",
    "ENGLISH 10",
)

PAIRED_COLUMNS = (
    ("Composition INT", "Composition DEU"),
    ("Variation INT", "Variation DEU"),
    ("Reconstruction INT", "Reconstruction DEU"),
    ("Source-2 INT", "Source-2 DEU"),
    ("Base INT", "Base DEU"),
)

EXPECTED_WORD_CLASSES = {
    "ADJ",
    "ADV",
    "ART",
    "CONJ",
    "INTERJ",
    "N",
    "NP",
    "NUM",
    "PREF",
    "PREFV",
    "PREP",
    "PRON",
    "PTCL",
    "PTCLV",
    "V",
    "VP",
}

EXPECTED_NOUN_FLEXION = {"M", "F", "M/F", "PL"}
PARADIGM_KEY_RE = re.compile(
    r"^(ADJ-DECL|EXIST|[A-Z]+(?:-[A-Za-z0-9∅]+)+|V-(?:IRR-)?\d+)$"
)

PARADIGM_ALIASES = {
    # The glossary and paradigm table use slightly different spellings.
    "NME-i": "NM-E-i",
    "NME-IRR-01": "NM-IRR-01",
    "NME-IRR-02": "NM-IRR-02",
    "NME-IRR-03": "NM-IRR-03",
    "NME-IRR-04": "NM-IRR-04",
    "NME-IRR-05": "NM-IRR-05",
    # Observed zero-ending spelling drift.
    "NFPE-∅-01": "NFPE-∅-1",
}

SOURCE_MARKER_LABELS = {
    "→": {
        "en": "internal derivation or word-family link",
        "de": "interne Ableitung oder Wortfamilienbezug",
    },
    "←": {
        "en": "derived from or borrowed from",
        "de": "abgeleitet oder entlehnt aus",
    },
    "etym?": {
        "en": "no confirmed etymology",
        "de": "keine gesicherte Etymologie",
    },
}

GRAMMAR_EXPLANATIONS = {
    "RECT": {
        "en": "A case group from the stakeholder tables. In the forms table it marks the direct/basic case row.",
        "de": "Fallgruppe aus den Tabellen des Stakeholders. In der Formentabelle markiert sie den direkten/Grund-Fall.",
    },
    "OBL": {
        "en": "A case group from the stakeholder tables for oblique forms, i.e. forms used with specific cases such as accusative, dative, locative, and others.",
        "de": "Fallgruppe aus den Tabellen des Stakeholders für oblique Formen, also Formen mit bestimmten Fällen wie Akkusativ, Dativ, Lokativ usw.",
    },
    "NOM": {
        "en": "Nominative. The abbreviation table glosses this as the 'who-case'.",
        "de": "Nominativ. Die Abkürzungstabelle erklärt ihn als Wer-Fall.",
    },
    "ACC": {
        "en": "Accusative. The abbreviation table glosses this as the 'whom-case'.",
        "de": "Akkusativ. Die Abkürzungstabelle erklärt ihn als Wen-Fall.",
    },
    "DAT": {
        "en": "Dative. The abbreviation table glosses this as the 'to/for whom-case'.",
        "de": "Dativ. Die Abkürzungstabelle erklärt ihn als Wem-Fall.",
    },
    "ABL": {
        "en": "Ablative. The abbreviation table glosses this as the 'from where-case'.",
        "de": "Ablativ. Die Abkürzungstabelle erklärt ihn als Woher-Fall.",
    },
    "LOC": {
        "en": "Locative. In the case tables this is the 'where/where-to-case'.",
        "de": "Lokativ. In den Kasustabellen ist das der Wo-/Wohin-Fall.",
    },
    "INS/SOC": {
        "en": "Instrumental/social case row from the stakeholder tables.",
        "de": "Instrumental-/Soziativ-Zeile aus den Tabellen des Stakeholders.",
    },
    "GEN": {
        "en": "Genitive. The abbreviation table glosses this as the 'whose-case'.",
        "de": "Genitiv. Die Abkürzungstabelle erklärt ihn als Wes(sen)-Fall.",
    },
    "SG": {"en": "Singular: one.", "de": "Singular: Einzahl."},
    "PL": {"en": "Plural: more than one.", "de": "Plural: Mehrzahl."},
    "M": {"en": "Masculine gender/class.", "de": "Maskuline Klasse."},
    "F": {"en": "Feminine gender/class.", "de": "Feminine Klasse."},
    "M/F": {"en": "Masculine/feminine class; the table contains both sets of forms.", "de": "Maskulin/feminine Klasse; die Tabelle enthält beide Formreihen."},
    "NPFV": {
        "en": "Non-perfective marker/category from the verb table.",
        "de": "Nichtperfektive Markierung/Kategorie aus der Verbtabelle.",
    },
    "PRS": {"en": "Present tense.", "de": "Präsens/Gegenwart."},
    "FUT": {"en": "Future tense.", "de": "Futur/Zukunft."},
    "PST": {"en": "Past tense.", "de": "Vergangenheit."},
    "NEG": {"en": "Negative form.", "de": "Negative/verneinte Form."},
    "1SG": {"en": "1st person singular.", "de": "1. Person Singular."},
    "2SG": {"en": "2nd person singular.", "de": "2. Person Singular."},
    "3SG": {"en": "3rd person singular.", "de": "3. Person Singular."},
    "1PL": {"en": "1st person plural.", "de": "1. Person Plural."},
    "2PL": {"en": "2nd person plural.", "de": "2. Person Plural."},
    "3PL": {"en": "3rd person plural.", "de": "3. Person Plural."},
}


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text in DASH_VALUES else text


def clean_display_lemma(value: str) -> str:
    return value.replace("-", "")


def stem_from_lemma(value: str) -> str:
    if "-" not in value:
        return clean_display_lemma(value)
    return value.split("-", 1)[0]


def combine_stem_and_form(stem: str, form: str) -> str:
    if not form:
        return stem
    if form.startswith("-"):
        return stem + form[1:]
    return form


def resolve_paradigm(paradigm: str) -> str:
    return PARADIGM_ALIASES.get(paradigm, paradigm)


def compact_dict(data: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in data.items() if value not in ("", [], {})}


def source_path(workbook: Path) -> str:
    try:
        return str(workbook.relative_to(ROOT))
    except ValueError:
        return str(workbook)


def first_present(*values: str) -> str:
    for value in values:
        cleaned = clean_value(value)
        if cleaned:
            return cleaned
    return ""


def value_list(row: pd.Series, prefix: str, count: int = 10) -> list[str]:
    values: list[str] = []
    for index in range(1, count + 1):
        value = clean_value(row.get(f"{prefix} {index:02d}", ""))
        if value:
            values.append(value)
    return values


def paired_detail(row: pd.Series, int_col: str, deu_col: str) -> dict[str, str]:
    return compact_dict(
        {
            "int": clean_value(row.get(int_col, "")),
            "deu": clean_value(row.get(deu_col, "")),
        }
    )


def entry_id(source_row: int, roman_int: str, roman_deu: str) -> str:
    seed = f"{source_row}|{roman_int}|{roman_deu}"
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:8]
    return f"g{source_row:05d}_{digest}"


def load_glossary(workbook: Path) -> pd.DataFrame:
    df = pd.read_excel(
        workbook,
        sheet_name=GLOSSARY_SHEET,
        dtype=str,
        keep_default_na=False,
    )
    df = df.map(clean_value)
    dropped_empty_columns = [
        column
        for column in df.columns
        if str(column).startswith("Unnamed:") and not df[column].any()
    ]
    if dropped_empty_columns:
        df = df.drop(columns=dropped_empty_columns)
    df.attrs["dropped_empty_columns"] = dropped_empty_columns
    return df


def validate_glossary_columns(df: pd.DataFrame) -> None:
    expected = list(EXPECTED_GLOSSARY_COLUMNS)
    actual = list(df.columns)
    missing = [column for column in expected if column not in actual]
    extra = [column for column in actual if column not in expected]
    if missing or extra:
        raise ValueError(
            "Unexpected GLOSSARY columns. "
            f"Missing: {missing or 'none'}. Extra: {extra or 'none'}."
        )


def load_glossary_hyperlinks(workbook: Path) -> dict[int, dict[str, str]]:
    """Return source links keyed by one-based workbook row."""
    wb = load_workbook(workbook, read_only=False, data_only=True)
    ws = wb[GLOSSARY_SHEET]
    headers = {
        clean_value(cell.value): cell.column
        for cell in ws[1]
        if clean_value(cell.value)
    }
    links: dict[int, dict[str, str]] = {}
    for source_column, output_key in (
        ("Source-2 INT", "source_2_int_url"),
        ("Source-2 DEU", "source_2_deu_url"),
    ):
        column = headers.get(source_column)
        if not column:
            continue
        for source_row in range(2, ws.max_row + 1):
            hyperlink = ws.cell(source_row, column).hyperlink
            if hyperlink and hyperlink.target:
                links.setdefault(source_row, {})[output_key] = hyperlink.target
    wb.close()
    return links


def build_entries(
    df: pd.DataFrame,
    source_hyperlinks: dict[int, dict[str, str]] | None = None,
) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    source_hyperlinks = source_hyperlinks or {}
    for zero_index, row in df.iterrows():
        source_row = int(zero_index) + 2
        roman_int = clean_value(row.get("ROMAN INT", ""))
        roman_deu = clean_value(row.get("ROMAN DEU", ""))

        details = compact_dict(
            {
                "composition": paired_detail(row, "Composition INT", "Composition DEU"),
                "variation": paired_detail(row, "Variation INT", "Variation DEU"),
                "reconstruction": paired_detail(
                    row,
                    "Reconstruction INT",
                    "Reconstruction DEU",
                ),
                "source": compact_dict(
                    {
                        "source_1": clean_value(row.get("Source-1", "")),
                        "source_2_int": clean_value(row.get("Source-2 INT", "")),
                        "source_2_deu": clean_value(row.get("Source-2 DEU", "")),
                        **source_hyperlinks.get(source_row, {}),
                    }
                ),
                "base": paired_detail(row, "Base INT", "Base DEU"),
            }
        )

        grammar = compact_dict(
            {
                "word_class_1": clean_value(row.get("Word class 1", "")),
                "word_class_2": clean_value(row.get("Word class 2", "")),
                "flexion_1": clean_value(row.get("Flexion 1", "")),
                "flexion_2_int": clean_value(row.get("Flexion 2 INT", "")),
                "flexion_2_deu": clean_value(row.get("Flexion 2 DEU", "")),
                "flexion_3_int": clean_value(row.get("Flexion 3 INT", "")),
                "flexion_3_deu": clean_value(row.get("Flexion 3 DEU", "")),
                "paradigm": clean_value(row.get("Paradigm", "")),
                "domain": clean_value(row.get("Domain", "")),
            }
        )

        entry = compact_dict(
            {
                "id": entry_id(source_row, roman_int, roman_deu),
                "source": {"sheet": GLOSSARY_SHEET, "row": source_row},
                "lemma": compact_dict(
                    {
                        "int": roman_int,
                        "deu": roman_deu,
                        "display_int": clean_display_lemma(roman_int),
                        "display_deu": clean_display_lemma(roman_deu),
                    }
                ),
                "grammar": grammar,
                "meanings": compact_dict(
                    {
                        "de": value_list(row, "DEUTSCH"),
                        "en": value_list(row, "ENGLISH"),
                    }
                ),
                "details": details,
            }
        )
        entries.append(entry)
    return entries


def build_search_entries(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    search_entries: list[dict[str, Any]] = []
    for entry in entries:
        lemma = entry.get("lemma", {})
        grammar = entry.get("grammar", {})
        meanings = entry.get("meanings", {})
        search_entries.append(
            compact_dict(
                {
                    "id": entry["id"],
                    "source_row": entry["source"]["row"],
                    "roman_int": lemma.get("display_int", lemma.get("int", "")),
                    "roman_deu": lemma.get("display_deu", lemma.get("deu", "")),
                    "raw_roman_int": lemma.get("int", ""),
                    "raw_roman_deu": lemma.get("deu", ""),
                    "word_class": grammar.get("word_class_1", ""),
                    "word_class_label": grammar.get("labels", {}).get("word_class_1", ""),
                    "subclass": grammar.get("word_class_2", ""),
                    "subclass_label": grammar.get("labels", {}).get("word_class_2", ""),
                    "de": meanings.get("de", [])[:3],
                    "en": meanings.get("en", [])[:3],
                }
            )
        )
    return search_entries


def sheet_to_records(workbook: Path, sheet_name: str) -> list[dict[str, str]]:
    df = pd.read_excel(workbook, sheet_name=sheet_name, dtype=str, keep_default_na=False)
    df = df.map(clean_value)
    return [
        compact_dict({str(key): clean_value(value) for key, value in row.items()})
        for row in df.to_dict(orient="records")
    ]


def sheet_to_matrix(workbook: Path, sheet_name: str) -> list[list[str]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb[sheet_name]
    matrix: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        matrix.append([clean_value(value) for value in row])

    while matrix and not any(matrix[-1]):
        matrix.pop()

    return matrix


def lexical_abbreviation_rows(workbook: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in sheet_to_matrix(workbook, "abbrs-lex"):
        item = compact_dict(
            {
                "de_code": clean_value(row[0] if len(row) > 0 else ""),
                "de": clean_value(row[1] if len(row) > 1 else ""),
                "en_code": clean_value(row[3] if len(row) > 3 else ""),
                "en": clean_value(row[4] if len(row) > 4 else ""),
            }
        )
        if item:
            rows.append(item)
    return rows


def extract_paradigm_keys(paradigm_tables: dict[str, list[list[str]]]) -> dict[str, list[str]]:
    keys: dict[str, set[str]] = {sheet: set() for sheet in paradigm_tables}
    keys["ADJ-DECL"].add("ADJ-DECL")
    keys["V-EXIST"].add("EXIST")

    for sheet, matrix in paradigm_tables.items():
        for row in matrix[:4]:
            for value in row:
                if PARADIGM_KEY_RE.match(value):
                    keys[sheet].add(value)

    return {sheet: sorted(values) for sheet, values in keys.items()}


def section_bounds(matrix: list[list[str]]) -> dict[str, tuple[int, int]]:
    markers: list[tuple[str, int]] = []
    for index, value in enumerate(matrix[0]):
        if value in {"INT", "DEU"}:
            markers.append((value.lower(), index))

    bounds: dict[str, tuple[int, int]] = {}
    for marker_index, (name, start) in enumerate(markers):
        end = markers[marker_index + 1][1] if marker_index + 1 < len(markers) else len(matrix[0])
        bounds[name] = (start, end)
    return bounds


def meaningful_data_rows(matrix: list[list[str]], start_index: int) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in matrix[start_index:]:
        if not any(row):
            if rows:
                break
            continue
        rows.append(row)
    return rows


def parse_declension_table(
    matrix: list[list[str]],
    *,
    has_gender_row: bool = False,
) -> dict[str, dict[str, list[dict[str, str]]]]:
    parsed: dict[str, dict[str, list[dict[str, str]]]] = {"int": {}, "deu": {}}
    bounds = section_bounds(matrix)
    header_row = matrix[1]
    gender_row = matrix[2] if has_gender_row else []
    data_start = 3 if has_gender_row else 2

    for spelling, (start, end) in bounds.items():
        current_paradigm = ""
        for column in range(start, end):
            header = clean_value(header_row[column] if column < len(header_row) else "")
            if header:
                current_paradigm = header
            if not current_paradigm:
                continue

            gender = clean_value(gender_row[column] if column < len(gender_row) else "")
            key = f"{current_paradigm}:{gender}" if gender else current_paradigm
            rows: list[dict[str, str]] = []
            for row in meaningful_data_rows(matrix, data_start):
                form = clean_value(row[column] if column < len(row) else "")
                rows.append(
                    compact_dict(
                        {
                            "case_group": clean_value(row[0] if len(row) > 0 else ""),
                            "case": clean_value(row[1] if len(row) > 1 else ""),
                            "number": clean_value(row[2] if len(row) > 2 else ""),
                            "gender": gender,
                            "ending": form,
                        }
                    )
                )
            parsed[spelling][key] = rows
    return parsed


def parse_adjective_table(matrix: list[list[str]]) -> dict[str, list[dict[str, str]]]:
    rows: list[dict[str, str]] = []
    for row in meaningful_data_rows(matrix, 2):
        for gender, column in (("M", 2), ("F", 3)):
            rows.append(
                compact_dict(
                    {
                        "case": clean_value(row[0] if len(row) > 0 else ""),
                        "number": clean_value(row[1] if len(row) > 1 else ""),
                        "gender": gender,
                        "ending": clean_value(row[column] if column < len(row) else ""),
                    }
                )
            )
    return {"ADJ-DECL": rows}


def parse_verb_table(matrix: list[list[str]]) -> dict[str, dict[str, list[dict[str, str]]]]:
    parsed: dict[str, dict[str, list[dict[str, str]]]] = {"int": {}, "deu": {}}
    bounds = section_bounds(matrix)
    header_row = matrix[1]

    for spelling, (start, end) in bounds.items():
        for column in range(start, end):
            paradigm = clean_value(header_row[column] if column < len(header_row) else "")
            if not paradigm:
                continue

            rows: list[dict[str, str]] = []
            for row in meaningful_data_rows(matrix, 2):
                ending = clean_value(row[column] if column < len(row) else "")
                rows.append(
                    compact_dict(
                        {
                            "aspect": clean_value(row[0] if len(row) > 0 else ""),
                            "tense": clean_value(row[1] if len(row) > 1 else ""),
                            "person_number": clean_value(row[2] if len(row) > 2 else ""),
                            "ending": ending,
                        }
                    )
                )
            parsed[spelling][paradigm] = rows
    return parsed


def parse_exist_table(matrix: list[list[str]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in meaningful_data_rows(matrix, 1):
        rows.append(
            compact_dict(
                {
                    "tense": clean_value(row[0] if len(row) > 0 else ""),
                    "person_number": clean_value(row[1] if len(row) > 1 else ""),
                    "polarity": clean_value(row[2] if len(row) > 2 else ""),
                    "form_int": clean_value(row[3] if len(row) > 3 else ""),
                    "form_deu": clean_value(row[3] if len(row) > 3 else ""),
                    "german": clean_value(row[4] if len(row) > 4 else ""),
                    "english": clean_value(row[5] if len(row) > 5 else ""),
                }
            )
        )
    return rows


def build_paradigm_model(paradigm_tables: dict[str, list[list[str]]]) -> dict[str, Any]:
    return {
        "adjectives": parse_adjective_table(paradigm_tables["ADJ-DECL"]),
        "nouns": {
            "M": parse_declension_table(paradigm_tables["M-DECL"]),
            "F": parse_declension_table(paradigm_tables["F-DECL"]),
            "M/F": parse_declension_table(paradigm_tables["MF-DECL"], has_gender_row=True),
        },
        "verbs": parse_verb_table(paradigm_tables["V-CONJG"]),
        "exist": parse_exist_table(paradigm_tables["V-EXIST"]),
    }


def add_form_labels(row: dict[str, str], references: dict[str, Any]) -> dict[str, str]:
    parts = [
        row.get("aspect", ""),
        row.get("tense", ""),
        row.get("case_group", ""),
        row.get("case", ""),
        row.get("number", ""),
        row.get("person_number", ""),
        row.get("gender", ""),
        row.get("polarity", ""),
    ]
    label_parts_en = [reference_label(references, part, "en") for part in parts if part]
    label_parts_de = [reference_label(references, part, "de") for part in parts if part]
    explanation_parts_en = [
        references.get("grammar_abbreviations", {}).get(part, {}).get("explanation_en", "")
        for part in parts
        if part
    ]
    explanation_parts_de = [
        references.get("grammar_abbreviations", {}).get(part, {}).get("explanation_de", "")
        for part in parts
        if part
    ]
    return {
        **row,
        "label_en": " · ".join(label_parts_en),
        "label_de": " · ".join(label_parts_de),
        "explanation_en": " ".join(part for part in explanation_parts_en if part),
        "explanation_de": " ".join(part for part in explanation_parts_de if part),
    }


def generate_adjective_forms(
    entry: dict[str, Any],
    paradigm_model: dict[str, Any],
    references: dict[str, Any],
) -> dict[str, Any]:
    paradigm = entry.get("grammar", {}).get("paradigm", "")
    rows = paradigm_model["adjectives"].get(paradigm)
    if not rows:
        return {}

    forms: dict[str, list[dict[str, str]]] = {"int": [], "deu": []}
    for spelling in ("int", "deu"):
        stem = stem_from_lemma(entry.get("lemma", {}).get(spelling, ""))
        for row in rows:
            forms[spelling].append(
                add_form_labels(
                    {
                        **row,
                        "form": combine_stem_and_form(stem, row.get("ending", "")),
                    },
                    references,
                )
            )
    return {
        "kind": "adjective_declension",
        "status": "generated",
        "paradigm": paradigm,
        "forms": forms,
    }


def generate_noun_forms(
    entry: dict[str, Any],
    paradigm_model: dict[str, Any],
    references: dict[str, Any],
) -> dict[str, Any]:
    grammar = entry.get("grammar", {})
    gender_class = grammar.get("flexion_1", "")
    source_paradigm = grammar.get("paradigm", "")
    paradigm = resolve_paradigm(source_paradigm)
    noun_tables = paradigm_model["nouns"].get(gender_class)
    if not noun_tables or not paradigm:
        return {}

    forms: dict[str, list[dict[str, str]]] = {"int": [], "deu": []}
    notes: list[str] = []
    for spelling in ("int", "deu"):
        stem = stem_from_lemma(entry.get("lemma", {}).get(spelling, ""))
        if gender_class == "M/F":
            rows = []
            for key in (f"{paradigm}:M", f"{paradigm}:F"):
                rows.extend(noun_tables[spelling].get(key, []))
        else:
            rows = noun_tables[spelling].get(paradigm, [])

        if not rows:
            notes.append(f"No {spelling.upper()} table rows found for paradigm {paradigm}.")
            continue

        for row in rows:
            forms[spelling].append(
                add_form_labels(
                    {
                        **row,
                        "form": combine_stem_and_form(stem, row.get("ending", "")),
                    },
                    references,
                )
            )

    if not forms["int"] and not forms["deu"]:
        return {}

    return compact_dict(
        {
            "kind": "noun_declension",
            "status": "generated" if not notes else "partial",
            "paradigm": paradigm,
            "source_paradigm": source_paradigm if source_paradigm != paradigm else "",
            "gender_class": gender_class,
            "forms": forms,
            "notes": notes,
        }
    )


def generate_verb_forms(
    entry: dict[str, Any],
    paradigm_model: dict[str, Any],
    references: dict[str, Any],
) -> dict[str, Any]:
    source_paradigm = entry.get("grammar", {}).get("paradigm", "")
    paradigm = resolve_paradigm(source_paradigm)
    if not paradigm:
        return {}

    if paradigm == "EXIST":
        rows = [
            add_form_labels(row, references)
            for row in paradigm_model["exist"]
        ]
        return {
            "kind": "verb_exist",
            "status": "generated",
            "paradigm": paradigm,
            "source_paradigm": source_paradigm if source_paradigm != paradigm else "",
            "forms": {"int": rows, "deu": rows},
        }

    forms: dict[str, list[dict[str, str]]] = {"int": [], "deu": []}
    notes: list[str] = []
    for spelling in ("int", "deu"):
        rows = paradigm_model["verbs"][spelling].get(paradigm, [])
        stem = stem_from_lemma(entry.get("lemma", {}).get(spelling, ""))
        if not rows:
            notes.append(f"No {spelling.upper()} verb rows found for paradigm {paradigm}.")
            continue
        for row in rows:
            forms[spelling].append(
                add_form_labels(
                    {
                        **row,
                        "form": combine_stem_and_form(stem, row.get("ending", "")),
                    },
                    references,
                )
            )

    if not forms["int"] and not forms["deu"]:
        return {}

    return compact_dict(
        {
            "kind": "verb_conjugation",
            "status": "generated" if not notes else "partial",
            "paradigm": paradigm,
            "source_paradigm": source_paradigm if source_paradigm != paradigm else "",
            "forms": forms,
            "notes": notes,
        }
    )


def enrich_entries(
    entries: list[dict[str, Any]],
    paradigm_model: dict[str, Any],
    references: dict[str, Any],
) -> list[dict[str, Any]]:
    for entry in entries:
        grammar = entry.get("grammar", {})
        word_class = grammar.get("word_class_1", "")
        source = entry.get("details", {}).get("source", {})

        grammar["labels"] = compact_dict(
            {
                "word_class_1": label_with_code(references, grammar.get("word_class_1", "")),
                "word_class_2": label_with_code(references, grammar.get("word_class_2", "")),
                "flexion_1": label_with_code(references, grammar.get("flexion_1", "")),
                "domain": label_with_code(references, grammar.get("domain", "")),
            }
        )
        source_1 = source.get("source_1", "")
        if source_1:
            source["source_1_label"] = label_with_code(references, source_1)

        morphology = morphology_status(entry, paradigm_model)
        if morphology:
            entry["morphology"] = morphology

    return entries


def morphology_status(entry: dict[str, Any], paradigm_model: dict[str, Any]) -> dict[str, Any]:
    grammar = entry.get("grammar", {})
    word_class = grammar.get("word_class_1", "")
    source_paradigm = grammar.get("paradigm", "")
    paradigm = resolve_paradigm(source_paradigm)
    flexion_1 = grammar.get("flexion_1", "")

    if word_class == "ADJ":
        return compact_dict(
            {
                "kind": "adjective_declension",
                "available": paradigm in paradigm_model["adjectives"],
                "paradigm": paradigm,
                "source_paradigm": source_paradigm if source_paradigm != paradigm else "",
            }
        )

    if word_class == "N":
        noun_tables = paradigm_model["nouns"].get(flexion_1)
        available = False
        if noun_tables and paradigm:
            if flexion_1 == "M/F":
                available = any(
                    noun_tables[spelling].get(f"{paradigm}:{gender}")
                    for spelling in ("int", "deu")
                    for gender in ("M", "F")
                )
            else:
                available = any(
                    noun_tables[spelling].get(paradigm)
                    for spelling in ("int", "deu")
                )
        return compact_dict(
            {
                "kind": "noun_declension",
                "available": available,
                "paradigm": paradigm,
                "source_paradigm": source_paradigm if source_paradigm != paradigm else "",
                "gender_class": flexion_1,
            }
        )

    if word_class == "V":
        available = paradigm == "EXIST" or any(
            paradigm_model["verbs"][spelling].get(paradigm)
            for spelling in ("int", "deu")
        )
        return compact_dict(
            {
                "kind": "verb_conjugation" if paradigm != "EXIST" else "verb_exist",
                "available": available,
                "paradigm": paradigm,
                "source_paradigm": source_paradigm if source_paradigm != paradigm else "",
            }
        )

    return {}


def structure_word_classes(workbook: Path) -> list[str]:
    try:
        df = pd.read_excel(workbook, sheet_name="structure", dtype=str, keep_default_na=False)
    except ValueError:
        return []

    df = df.map(clean_value)
    if "Wordclass-1" not in df.columns:
        return []

    values = {
        value
        for value in df["Wordclass-1"].tolist()
        if value and value not in {"TYPE", "Wordclass-1"}
    }
    return sorted(values)


def build_references(workbook: Path) -> dict[str, Any]:
    grammar_rows = sheet_to_matrix(workbook, "abbrs-gram")
    language_rows = sheet_to_matrix(workbook, "abbrs-lang")
    lexical_rows = lexical_abbreviation_rows(workbook)

    grammar: dict[str, dict[str, str]] = {}
    grammar_variants: dict[str, list[dict[str, str]]] = {}
    for row in grammar_rows[1:]:
        code = clean_value(row[0] if len(row) > 0 else "")
        if not code:
            continue
        reference = compact_dict(
            {
                "en": clean_value(row[1] if len(row) > 1 else ""),
                "de": clean_value(row[2] if len(row) > 2 else ""),
                "de_plain": clean_value(row[3] if len(row) > 3 else ""),
                "explanation_en": GRAMMAR_EXPLANATIONS.get(code, {}).get("en", ""),
                "explanation_de": GRAMMAR_EXPLANATIONS.get(code, {}).get("de", ""),
                "roman_int": first_present(
                    row[4] if len(row) > 4 else "",
                    row[5] if len(row) > 5 else "",
                ),
                "roman_deu": first_present(
                    row[6] if len(row) > 6 else "",
                    row[7] if len(row) > 7 else "",
                ),
            }
        )
        grammar_variants.setdefault(code, []).append(reference)
        grammar[code] = reference

    languages: dict[str, dict[str, str]] = {}
    for row in language_rows[1:]:
        code = clean_value(row[0] if len(row) > 0 else "")
        if not code:
            continue
        languages[code] = compact_dict(
            {
                "iso": clean_value(row[1] if len(row) > 1 else ""),
                "en": clean_value(row[2] if len(row) > 2 else ""),
                "de": clean_value(row[3] if len(row) > 3 else ""),
                "roman_int": clean_value(row[4] if len(row) > 4 else ""),
                "roman_deu": clean_value(row[5] if len(row) > 5 else ""),
            }
        )

    lexical: dict[str, dict[str, str]] = {"de": {}, "en": {}}
    for row in lexical_rows:
        de_code = row.get("de_code", "")
        en_code = row.get("en_code", "")
        if de_code:
            lexical["de"][de_code] = row.get("de", "")
        if en_code:
            lexical["en"][en_code] = row.get("en", "")

    return {
        "grammar_abbreviations": grammar,
        "grammar_abbreviation_variants": {
            code: variants
            for code, variants in grammar_variants.items()
            if len(variants) > 1
        },
        "language_abbreviations": languages,
        "lexical_abbreviations": lexical,
        "source_markers": SOURCE_MARKER_LABELS,
    }


def reference_label(references: dict[str, Any], code: str, language: str = "en") -> str:
    code = clean_value(code)
    if not code:
        return ""
    grammar = references.get("grammar_abbreviations", {}).get(code, {})
    source_marker = references.get("source_markers", {}).get(code, {})
    language_ref = references.get("language_abbreviations", {}).get(code, {})
    lexical_ref = references.get("lexical_abbreviations", {}).get(language, {}).get(code, "")
    return first_present(
        grammar.get(language, ""),
        source_marker.get(language, ""),
        language_ref.get(language, ""),
        lexical_ref,
        code,
    )


def label_with_code(references: dict[str, Any], code: str, language: str = "en") -> str:
    label = reference_label(references, code, language)
    if not code:
        return ""
    if not label or label == code:
        return code
    return f"{code} · {label}"


def issue(
    code: str,
    severity: str,
    message: str,
    row: int | None = None,
    entry_id_value: str | None = None,
    lemma: str | None = None,
    fields: list[str] | None = None,
    values: dict[str, str] | None = None,
) -> dict[str, Any]:
    data: dict[str, Any] = {
        "code": code,
        "severity": severity,
        "message": message,
    }
    if row is not None:
        data["row"] = row
    if entry_id_value:
        data["entry_id"] = entry_id_value
    if lemma:
        data["lemma"] = lemma
    if fields:
        data["fields"] = fields
    if values:
        data["values"] = values
    return data


def build_validation_report(
    workbook: Path,
    df: pd.DataFrame,
    entries: list[dict[str, Any]],
    paradigm_keys: dict[str, list[str]],
) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    entries_by_row = {entry["source"]["row"]: entry for entry in entries}
    structure_classes = structure_word_classes(workbook)
    observed_classes = sorted({value for value in df["Word class 1"].tolist() if value})
    references = build_references(workbook)

    for column in df.attrs.get("dropped_empty_columns", []):
        issues.append(
            issue(
                "empty_formatting_column_ignored",
                "info",
                "An unnamed, fully empty Excel formatting column was ignored.",
                fields=[str(column)],
            )
        )

    for code, variants in references.get("grammar_abbreviation_variants", {}).items():
        issues.append(
            issue(
                "duplicate_grammar_abbreviation",
                "warning",
                "Grammar abbreviation has multiple source definitions; all variants are preserved.",
                fields=["abbrs-gram"],
                values={"code": code, "variant_count": str(len(variants))},
            )
        )

    lexical_masculine = references.get("lexical_abbreviations", {}).get("en", {}).get("m.", "")
    if lexical_masculine.lower() == "mackuline":
        issues.append(
            issue(
                "suspected_reference_typo",
                "warning",
                "Lexical abbreviation label appears to contain a source typo.",
                fields=["abbrs-lex", "English m."],
                values={"source_value": lexical_masculine},
            )
        )

    all_paradigm_keys = {
        key
        for keys in paradigm_keys.values()
        for key in keys
    }

    for row_index, row in df.iterrows():
        source_row = int(row_index) + 2
        entry = entries_by_row[source_row]
        entry_id_value = entry["id"]
        lemma = entry.get("lemma", {}).get("display_int", "")
        word_class = clean_value(row.get("Word class 1", ""))
        flexion_1 = clean_value(row.get("Flexion 1", ""))
        paradigm = clean_value(row.get("Paradigm", ""))
        german_meanings = value_list(row, "DEUTSCH")
        english_meanings = value_list(row, "ENGLISH")

        if not clean_value(row.get("ROMAN INT", "")):
            issues.append(
                issue(
                    "missing_roman_int",
                    "error",
                    "Entry has no Roman INT lemma.",
                    source_row,
                    entry_id_value,
                    lemma,
                    ["ROMAN INT"],
                )
            )
        if not clean_value(row.get("ROMAN DEU", "")):
            issues.append(
                issue(
                    "missing_roman_deu",
                    "error",
                    "Entry has no Roman DEU lemma.",
                    source_row,
                    entry_id_value,
                    lemma,
                    ["ROMAN DEU"],
                )
            )
        if not german_meanings:
            issues.append(
                issue(
                    "missing_german_meaning",
                    "error",
                    "Entry has no German meaning.",
                    source_row,
                    entry_id_value,
                    lemma,
                    [f"DEUTSCH {index:02d}" for index in range(1, 11)],
                )
            )
        if not english_meanings:
            issues.append(
                issue(
                    "missing_english_meaning",
                    "info",
                    "Entry has no English meaning yet.",
                    source_row,
                    entry_id_value,
                    lemma,
                    [f"ENGLISH {index:02d}" for index in range(1, 11)],
                )
            )

        if word_class not in EXPECTED_WORD_CLASSES:
            issues.append(
                issue(
                    "unexpected_word_class",
                    "error",
                    "Word class is not part of the currently expected dictionary classes.",
                    source_row,
                    entry_id_value,
                    lemma,
                    ["Word class 1"],
                    {"Word class 1": word_class},
                )
            )

        for int_col, deu_col in PAIRED_COLUMNS:
            int_value = clean_value(row.get(int_col, ""))
            deu_value = clean_value(row.get(deu_col, ""))
            if bool(int_value) != bool(deu_value):
                issues.append(
                    issue(
                        "paired_int_deu_asymmetry",
                        "warning",
                        "Paired INT/DEU fields should usually both be filled or both be empty.",
                        source_row,
                        entry_id_value,
                        lemma,
                        [int_col, deu_col],
                        {int_col: int_value, deu_col: deu_value},
                    )
                )

        if word_class == "N":
            if flexion_1 not in EXPECTED_NOUN_FLEXION:
                issues.append(
                    issue(
                        "unexpected_noun_flexion",
                        "warning",
                        "Noun has an unexpected Flexion 1 value.",
                        source_row,
                        entry_id_value,
                        lemma,
                        ["Word class 1", "Flexion 1"],
                        {"Word class 1": word_class, "Flexion 1": flexion_1},
                    )
                )
            if not paradigm:
                issues.append(
                    issue(
                        "missing_inflection_paradigm",
                        "warning",
                        "Noun has no paradigm, so forms cannot be generated automatically.",
                        source_row,
                        entry_id_value,
                        lemma,
                        ["Word class 1", "Paradigm"],
                        {"Word class 1": word_class, "Paradigm": paradigm},
                    )
                )
        elif word_class == "ADJ":
            if paradigm and paradigm != "ADJ-DECL":
                issues.append(
                    issue(
                        "unexpected_adjective_paradigm",
                        "warning",
                        "Adjective has a paradigm other than ADJ-DECL.",
                        source_row,
                        entry_id_value,
                        lemma,
                        ["Word class 1", "Paradigm"],
                        {"Word class 1": word_class, "Paradigm": paradigm},
                    )
                )
        elif word_class == "V":
            if not paradigm:
                issues.append(
                    issue(
                        "missing_inflection_paradigm",
                        "warning",
                        "Verb has no paradigm, so forms cannot be generated automatically.",
                        source_row,
                        entry_id_value,
                        lemma,
                        ["Word class 1", "Paradigm"],
                        {"Word class 1": word_class, "Paradigm": paradigm},
                    )
                )

        resolved_paradigm = resolve_paradigm(paradigm)
        if paradigm and resolved_paradigm != paradigm:
            issues.append(
                issue(
                    "paradigm_alias_applied",
                    "info",
                    "Glossary paradigm is resolved through an explicit alias before form generation.",
                    source_row,
                    entry_id_value,
                    lemma,
                    ["Paradigm"],
                    {"Paradigm": paradigm, "Resolved paradigm": resolved_paradigm},
                )
            )

        if paradigm and resolved_paradigm not in all_paradigm_keys:
            issues.append(
                issue(
                    "unmatched_paradigm_key",
                    "warning",
                    "Glossary references a paradigm key that was not found in the parsed paradigm table headers.",
                    source_row,
                    entry_id_value,
                    lemma,
                    ["Word class 1", "Flexion 1", "Paradigm"],
                    {
                        "Word class 1": word_class,
                        "Flexion 1": flexion_1,
                        "Paradigm": paradigm,
                        "Resolved paradigm": resolved_paradigm,
                    },
                )
            )

    if structure_classes:
        for value in observed_classes:
            if value not in structure_classes:
                issues.append(
                    issue(
                        "word_class_missing_from_structure_sheet",
                        "info",
                        "Observed Word class 1 value is not listed in the stakeholder structure sheet.",
                        None,
                        None,
                        None,
                        ["structure.Wordclass-1", "GLOSSARY.Word class 1"],
                        {"Word class 1": value},
                    )
                )
        for value in structure_classes:
            if value not in observed_classes:
                issues.append(
                    issue(
                        "structure_word_class_unused",
                        "info",
                        "Stakeholder structure sheet lists a Wordclass-1 value not currently used in GLOSSARY.",
                        None,
                        None,
                        None,
                        ["structure.Wordclass-1", "GLOSSARY.Word class 1"],
                        {"Wordclass-1": value},
                    )
                )

    counts_by_severity = Counter(item["severity"] for item in issues)
    counts_by_code = Counter(item["code"] for item in issues)
    issue_samples: dict[str, list[dict[str, Any]]] = {}
    for item in issues:
        code = item["code"]
        issue_samples.setdefault(code, [])
        if len(issue_samples[code]) < 12:
            issue_samples[code].append(item)

    return {
        "source_workbook": source_path(workbook),
        "issue_count": len(issues),
        "counts_by_severity": dict(sorted(counts_by_severity.items())),
        "counts_by_code": dict(sorted(counts_by_code.items())),
        "observed_word_classes": observed_classes,
        "structure_sheet_word_classes": structure_classes,
        "parsed_paradigm_keys": paradigm_keys,
        "paradigm_aliases": PARADIGM_ALIASES,
        "dropped_empty_columns": df.attrs.get("dropped_empty_columns", []),
        "issue_samples": issue_samples,
        "issues": issues,
    }


def build_summary(workbook: Path, df: pd.DataFrame, entries: list[dict[str, Any]]) -> dict[str, Any]:
    de_count = sum(1 for entry in entries if entry.get("meanings", {}).get("de"))
    en_count = sum(1 for entry in entries if entry.get("meanings", {}).get("en"))
    generated_forms_count = sum(
        1 for entry in entries if entry.get("morphology", {}).get("available")
    )
    word_class_counts = (
        df["Word class 1"].value_counts(dropna=False).rename_axis("word_class").to_dict()
    )
    paradigm_counts = df["Paradigm"].value_counts(dropna=False).rename_axis("paradigm").to_dict()
    source_hyperlink_count = sum(
        1
        for entry in entries
        for key in entry.get("details", {}).get("source", {})
        if key.endswith("_url")
    )

    return {
        "source_workbook": source_path(workbook),
        "glossary_entries": len(entries),
        "entries_with_german_meaning": de_count,
        "entries_with_english_meaning": en_count,
        "entries_with_generated_forms": generated_forms_count,
        "preserved_source_hyperlinks": source_hyperlink_count,
        "ignored_empty_formatting_columns": df.attrs.get("dropped_empty_columns", []),
        "word_class_counts": word_class_counts,
        "top_paradigm_counts": dict(list(paradigm_counts.items())[:30]),
        "output_files": [
            "entries.json",
            "entries_search.json",
            "abbreviations.json",
            "references.json",
            "paradigm_model.json",
            "paradigm_tables.json",
            "reports/data_coverage.json",
            "reports/validation_report.json",
            "reports/validation_summary.json",
            "summary.json",
        ],
    }


def build_data_coverage_report(
    workbook: Path,
    df: pd.DataFrame,
    entries: list[dict[str, Any]],
) -> dict[str, Any]:
    column_mapping = {
        "ROMAN INT": "entries[].lemma.int and entries_search[].raw_roman_int",
        "ROMAN DEU": "entries[].lemma.deu and entries_search[].raw_roman_deu",
        "Composition INT": "entries[].details.composition.int",
        "Composition DEU": "entries[].details.composition.deu",
        "Variation INT": "entries[].details.variation.int",
        "Variation DEU": "entries[].details.variation.deu",
        "Reconstruction INT": "entries[].details.reconstruction.int",
        "Reconstruction DEU": "entries[].details.reconstruction.deu",
        "Source-1": "entries[].details.source.source_1 and source_1_label",
        "Source-2 INT": "entries[].details.source.source_2_int and source_2_int_url when hyperlinked",
        "Source-2 DEU": "entries[].details.source.source_2_deu and source_2_deu_url when hyperlinked",
        "Base INT": "entries[].details.base.int and viewer word-family index",
        "Base DEU": "entries[].details.base.deu and viewer word-family index",
        "Word class 1": "entries[].grammar.word_class_1 and readable labels",
        "Word class 2": "entries[].grammar.word_class_2 and readable labels",
        "Flexion 1": "entries[].grammar.flexion_1 and morphology.gender_class where applicable",
        "Flexion 2 INT": "entries[].grammar.flexion_2_int",
        "Flexion 2 DEU": "entries[].grammar.flexion_2_deu",
        "Flexion 3 INT": "entries[].grammar.flexion_3_int",
        "Flexion 3 DEU": "entries[].grammar.flexion_3_deu",
        "Paradigm": "entries[].grammar.paradigm, entries[].morphology, and generated forms",
        "Domain": "entries[].grammar.domain; kept in Details because PDF marks it internal",
    }
    for index in range(1, 11):
        column_mapping[f"DEUTSCH {index:02d}"] = "entries[].meanings.de"
        column_mapping[f"ENGLISH {index:02d}"] = "entries[].meanings.en"

    non_empty_counts = {
        column: int((df[column].astype(str).str.strip() != "").sum())
        for column in df.columns
    }
    detail_counts = {
        key: sum(1 for entry in entries if entry.get("details", {}).get(key))
        for key in ("composition", "variation", "reconstruction", "source", "base")
    }
    morphology_counts = Counter()
    for entry in entries:
        morphology = entry.get("morphology", {})
        key = f"{morphology.get('kind', 'none')}:{morphology.get('available', False)}"
        morphology_counts[key] += 1

    return {
        "source_workbook": source_path(workbook),
        "glossary_columns": list(df.columns),
        "all_columns_mapped": set(df.columns) == set(column_mapping),
        "column_mapping": column_mapping,
        "non_empty_counts": non_empty_counts,
        "detail_counts": detail_counts,
        "morphology_counts": dict(sorted(morphology_counts.items())),
        "notes": [
            "Raw source workbook and PDF are preserved unchanged.",
            "Hyphenated lemmas preserve raw source values and add display_* values without internal hyphen markers.",
            "Paradigm and Domain are preserved, but treated as internal fields in the viewer per the PDF.",
            "Explicit paradigm aliases preserve the source value and add a resolved value only for form generation.",
            "Generated morphology is provisional until representative entries are reviewed by the linguistic stakeholder.",
        ],
    }


def write_json(path: Path, data: Any) -> None:
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def preprocess(workbook: Path, output_dir: Path) -> None:
    if not workbook.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook}")

    output_dir.mkdir(parents=True, exist_ok=True)
    reports_dir = output_dir / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    glossary = load_glossary(workbook)
    validate_glossary_columns(glossary)

    abbreviations = {
        "abbrs-gram": sheet_to_records(workbook, "abbrs-gram"),
        "abbrs-lang": sheet_to_records(workbook, "abbrs-lang"),
        "abbrs-lex": lexical_abbreviation_rows(workbook),
    }
    references = build_references(workbook)
    paradigm_tables = {sheet: sheet_to_matrix(workbook, sheet) for sheet in PARADIGM_SHEETS}
    paradigm_model = build_paradigm_model(paradigm_tables)
    source_hyperlinks = load_glossary_hyperlinks(workbook)
    entries = enrich_entries(
        build_entries(glossary, source_hyperlinks),
        paradigm_model,
        references,
    )
    search_entries = build_search_entries(entries)
    paradigm_keys = extract_paradigm_keys(paradigm_tables)
    validation_report = build_validation_report(workbook, glossary, entries, paradigm_keys)
    validation_summary = {
        key: validation_report[key]
        for key in (
            "source_workbook",
            "issue_count",
            "counts_by_severity",
            "counts_by_code",
            "observed_word_classes",
            "structure_sheet_word_classes",
            "parsed_paradigm_keys",
            "paradigm_aliases",
            "dropped_empty_columns",
            "issue_samples",
        )
    }
    summary = build_summary(workbook, glossary, entries)
    coverage_report = build_data_coverage_report(workbook, glossary, entries)

    write_json(output_dir / "entries.json", entries)
    write_json(output_dir / "entries_search.json", search_entries)
    write_json(output_dir / "abbreviations.json", abbreviations)
    write_json(output_dir / "references.json", references)
    write_json(output_dir / "paradigm_model.json", paradigm_model)
    write_json(output_dir / "paradigm_tables.json", paradigm_tables)
    write_json(reports_dir / "data_coverage.json", coverage_report)
    write_json(reports_dir / "validation_report.json", validation_report)
    write_json(reports_dir / "validation_summary.json", validation_summary)
    write_json(output_dir / "summary.json", summary)

    print(f"Wrote {len(entries):,} entries to {output_dir}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Source workbook path. Defaults to {DEFAULT_WORKBOOK}",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Output directory. Defaults to {DEFAULT_OUTPUT_DIR}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    preprocess(args.workbook.resolve(), args.out.resolve())


if __name__ == "__main__":
    main()
